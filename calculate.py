import openpyxl as xl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font 
import os


class Excelrw:
    def __init__(self,fpath):
        self.fpath =fpath
        self.cla1_fill =PatternFill('solid', fgColor="FFFF00")
        self.cla_1 =""
        self.cla_2 =""
        self.cla_data =""
        self.cla_unit =""
        self.cla_multi =""
        self.multi_mean_1 =""
        self.cla_sheet =""
        self.open_ok =0
        self.cla_head =""
        self.need_units =1


    def open_book(self):
        while(not self.open_ok):
            self.fpath=input("输入文件地址 \n")
            if not os.path.exists(self.fpath):
                print("文件地址错误")
            else:
                if not self.fpath.endswith(".xlsx" or ".XLSX"):
                    print("不是XLSX文件.如果是XLS请先转换为XLSX.")
                elif self.fpath.endswith(".xlsx" or ".XLSX"):
                    self._openbook()
                    self.open_ok=1

        
    def _openbook(self):
        print(f'加载 {self.fpath}')
        self.wb1 = load_workbook(self.fpath)
        self._set()
        self._cal()
        self._genxlsx()


    def _set(self):
        self.sheets = self.wb1.sheetnames
        print(self.sheets,end="")
        tmp_flag=1
        while(self.cla_sheet=="" and tmp_flag==1):
            self.cla_sheet = input("这是此文件中所含的数据表,\n请选择要分析的表格,用1,2,3...表示,或回车以选择默认值1: ")
            if self.cla_sheet=="":
                self.sheet_mean_1=input("使用默认值(即第1张表格)? (输入Y表示yes) ")
                if self.sheet_mean_1=="y" or self.sheet_mean_1=="Y":
                    self.cla_sheet="1"
                    tmp_flag=0
        self.wb=self.wb1[self.sheets[int(self.cla_sheet)-1]]
        while(self.cla_1==""):
            self.cla_1 = input("选择第一级分类(项目分部)所在的列: ")
        while(self.cla_2==""):
            self.cla_2 = input("选择第二级分类(项目名称)所在的列: ")
        while(self.cla_data==""):
            self.cla_data = input("选择单体工程量所在的列: ")
        tmp_flag=1
        while(self.cla_multi==""and tmp_flag==1):
            self.cla_multi = input("选择单体个数所在的列,或回车以选择将所有单体个数设置为1: ")
            if self.cla_multi=="":
                self.multi_mean_1=input("确定将所有单体个数设置为1? (输入Y表示yes) ")
                if self.multi_mean_1=="y" or self.multi_mean_1=="Y":
                    tmp_flag=0
        while(self.cla_unit==""):
            self.cla_unit = input("选择单位所在的列: ")
        while(self.cla_head==""):
            self.cla_head = input("选择数据开始的行号: ")
        tmp_foot = self.wb.max_row
        tmp_foot_ok=""
        while(tmp_foot_ok==""):
            tmp_foot_ok = input(f"自动检测数据的最后一行为{tmp_foot},输入Y表示认可,或输入行号来修改: ")
            if tmp_foot_ok=="Y" or tmp_foot_ok=="y":
                self.cla_foot = tmp_foot
            else :
                self.cla_foot = tmp_foot_ok
        # tmp_flag=1
        # while(tmp_flag==1):
        #     tmp_flag2 = input("是否需要保存其他的列?输入列号以选择或者直接回车表示不需要 ")
        
        #self.cla_unit = input("选择其他要保存的内容 ")
    
    def _cal(self):
        self.category={ }
        self.cla1_category={ }
        rows = range(int(self.cla_head),int(self.cla_foot)+1)
        print(f"将读取{self.cla_data}列{self.cla_head}到{self.cla_foot}行的数据")
        # print("项目名称大致如下: ",end="")
        # for row in rows:
        #    print(f"{self.wb[self.cla_2+str(row)].value} ",end="")
        print('数据处理中...')
        uncommon_rows=[]
        for row in rows:
            tmpc2=self.wb[self.cla_2+str(row)].value
            tmpc1=self.wb[self.cla_1+str(row)].value
            tmpunit=self.wb[self.cla_unit+str(row)].value
            if(tmpc1 is None or tmpc2 is None):
                uncommon_rows.append(row)
                continue
            if self.multi_mean_1=="Y" or self.multi_mean_1=="y":
                tmpv=self.wb[self.cla_data+str(row)].value
            else:
                tmpv=self.wb[self.cla_data+str(row)].value*self.wb[self.cla_multi+str(row)].value
            if tmpc1 not in self.category.keys():
                self.category.update({tmpc1:{}})
            if tmpc2 not in self.category[tmpc1].keys():
                tmpcate={'val':tmpv,'row':row,'unit':tmpunit}#后面可以通过行来补更多参数
                self.category[tmpc1].update({tmpc2:tmpcate})
            elif tmpc2 in self.category[tmpc1].keys():
                self.category[tmpc1][tmpc2]['val']+=tmpv
        #print(f"The result is {self.category}")
        print(f"第{uncommon_rows}行好像不是常规的行，已跳过")
        print("数据处理完成.")

    def _genxlsx(self):
        print("生成数据表中...")
        if 'Result' in self.wb1.sheetnames:
            del self.wb1['Result']
        self.wb1.create_sheet('Result')
        self.wbok=self.wb1['Result']
        ir=5
        self.wbok['C4']='工程项目'
        self.wbok['D4']='项目工程量'
        self.wbok['E4']='分部总工程量'
        self.wbok['F4']='单位'
        for cla1 in self.category:
            self.wbok['C'+str(ir)]=cla1
            self.wbok.cell(row=ir,column=3).fill=self.cla1_fill
            tmpir=ir
            tmpval=0
            ir+=1
            for cla2 in self.category[cla1]:
                self.wbok['C'+str(ir)]=cla2
                self.wbok['D'+str(ir)]=self.category[cla1][cla2]['val']
                self.wbok['F'+str(ir)]=self.category[cla1][cla2]['unit']
                tmpval+=self.category[cla1][cla2]['val']
                ir+=1
            self.wbok['E'+str(tmpir)]=tmpval
        print("生成完毕,可在原xlsx文件的Result数据表中查看.")
        try:
            self.wb1.save(self.fpath)
        except PermissionError:
            print("文件被占用,把excel关掉试试.")
            

if __name__ == '__main__':
    excelrw=Excelrw("")
    excelrw.open_book()