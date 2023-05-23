import tkinter as tk
from tkinter import *
import openpyxl as xl
from openpyxl import load_workbook
from tkinter import filedialog
from tkinter import ttk
from tkinter.messagebox import *
from calculate import Excelrw

class ShowExcel():
    def __init__(self,fpath):
        self.fpath=fpath
        self.openwin=[]
        self.win_openxls()
        print('OK')

    def win_openxls(self):
        print(self.fpath)
        self.wino=tk.Tk()
        self.wino.update()
        self.wino.destroy()
        self.w=xl.load_workbook(self.fpath)
        name=self.w.sheetnames
        print(name)

        sheet=self.w[name[0]]
        n=sheet.max_row
        col=sheet.max_column

        xlarr=[]
        for r in sheet.rows:
            arr=[]
            for cell in r:
                arr.append(cell.value)
            xlarr.append(arr)
    
        win=tk.Tk()
        win.geometry('800x500')

        colname=[]
        for c in range(col):
            colname.append(c)

        self.tree=ttk.Treeview(win,show='headings',columns=colname,selectmode = 'browse')#单行选中模式
        
        for c in range(col):
            self.tree.column(c,width=90,anchor='center')
            self.tree.heading(c,text=xlarr[0][c])#显示标题

        self.titles=xlarr[0]

        del(xlarr[0])#删了第一个,不删显示内容第一行与标题重复

        for i in range(n-1):
            self.tree.insert('',i,values=xlarr[i])#显示内容

        self.tree.pack(side=tk.TOP,fill=None,)

        btn1=tk.Button(win,text='保存', command=self.savebook)
        btn1.pack(side=tk.BOTTOM,expand = tk.YES)

        exceldd=Excelrw(self.fpath)
        btn2=tk.Button(win,text='计算',command=exceldd._openbook)
        btn2.pack(side=tk.BOTTOM,expand = tk.YES)

        self.tree.bind('<Double-Button-1>',self.viewclick)#后面用到,监控鼠标双击
        win.mainloop()

    def getv(self):

        editxt=self.enty.get()
        self.tree.set(self.sitem,(self.colint-1),editxt)
        self.openwin=[]
        self.nwin.destroy()

    def savebook(self):        
        ws=self.w.create_sheet('change1')
        ws.append(self.titles)
        for itm in self.tree.get_children():
            ws.append(self.tree.item(itm)['values'])   
        self.w.save(self.fpath)
        showinfo('提示','保存成功')


    def viewclick(self,event):

        for item in self.tree.selection():
            ttext=self.tree.item(item,'values')
            self.sitem=item

        col=self.tree.identify_column(event.x)
        self.colint=int(str(col.replace('#','')))
        
        self.nwin=tk.Tk()#编辑窗口
        self.nwin.geometry("360x200")
        label1 = tk.Label(self.nwin, text="修改：")
        label1.pack(side=tk.LEFT, fill=None)
        self.enty=tk.Text(self.nwin,width=300,height=300,wrap = tk.WORD)
        self.enty=tk.Entry(self.nwin)
        self.enty.pack(side=tk.LEFT, fill=None)
        btn=tk.Button(self.nwin,text='确认', command=self.getv)
        btn.pack(side=tk.LEFT,padx=6,ipadx=6)
        self.enty.insert('end',ttext[self.colint-1])#编辑框显示值
        self.openwin.append(self.nwin)
    
        if len(self.openwin)>1:
            self.openwin.pop(0).destroy()
        self.nwin.protocol('WM_DELETE_WINDOW', self.openwin)#绑定时件,关闭窗清除变量值
        self.nwin.mainloop()

    def show(self):
        print(self.fpath)
        wb = load_workbook(self.fpath)#载入excel
        print(wb.sheetnames)
        self.win_openxls

if __name__ == '__main__':
    path=input('读入地址')
    ds=ShowExcel(path)
    # ds.mainloop()

